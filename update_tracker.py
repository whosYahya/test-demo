"""
update_tracker.py
-----------------
Reads Playwright's results.json output and writes Pass/Fail results
into the next available Run column in AMC_Master_Tracker.xlsx.

Also updates the Test Case Register Status column.

Usage:
    python update_tracker.py --results test-results/results.json --excel AMC_Master_Tracker.xlsx

In GitHub Actions, add this step after your Playwright run:
    - name: Update QA Tracker
      run: python update_tracker.py --results test-results/results.json --excel AMC_Master_Tracker.xlsx
"""

import json
import argparse
import re
from datetime import date
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ── Colour constants ──────────────────────────────────────────────────────────
GREEN  = PatternFill('solid', start_color='C6EFCE')
RED    = PatternFill('solid', start_color='FFC7CE')
ORANGE = PatternFill('solid', start_color='FFEB9C')
GREY   = PatternFill('solid', start_color='D9D9D9')

GREEN_FONT  = Font(color='276221', bold=True)
RED_FONT    = Font(color='9C0006', bold=True)
ORANGE_FONT = Font(color='9C6500', bold=True)

# ── Run columns in Run Tracker (E to J = cols 5-10, 1-indexed) ───────────────
RUN_COL_START = 6   # Excel column index for "Run 1" (col F, row 3 is header)
RUN_COL_END   = 11  # Excel column index for "Run 6" (col K)
RUN_LABELS    = ['Run 1', 'Run 2', 'Run 3', 'Run 4', 'Run 5', 'Run 6']

# ── Header row index (1-based) ────────────────────────────────────────────────
HEADER_ROW    = 3   # Row 3 in both sheets contains column headers


def extract_tc_id(test_title: str) -> str | None:
    """
    Extract TC ID from Playwright test title.
    Supports formats:
      - "TC-HR-LEA-001 | ..."       → LEA-001
      - "LEA-001 | ..."             → LEA-001
      - "CUST-001 | ..."            → CUST-001
      - "TC-HR-CASE-001 | ..."      → CASE-001 (fallback)
    """
    # Match any XXX-NNN pattern at the start of the title
    match = re.search(r'\b([A-Z]{2,6}-\d{3})\b', test_title)
    return match.group(1) if match else None


def parse_playwright_results(results_path: str) -> dict:
    """
    Parse Playwright results.json.
    Returns dict: { tc_id: 'Pass' | 'Fail' | 'Skipped' }
    """
    with open(results_path, 'r') as f:
        data = json.load(f)

    results = {}

    for suite in data.get('suites', []):
        for spec in suite.get('specs', []):
            title = spec.get('title', '')
            tc_id = extract_tc_id(title)
            if not tc_id:
                continue

            # Playwright status: 'expected' = pass, 'unexpected' = fail, 'skipped'
            status_raw = spec.get('ok', True)
            tests = spec.get('tests', [])

            if tests:
                outcomes = [t.get('status', '') for t in tests]
                if any(o == 'skipped' for o in outcomes):
                    status = 'Skipped'
                elif all(o == 'expected' for o in outcomes):
                    status = 'Pass'
                else:
                    status = 'Fail'
            else:
                status = 'Pass' if status_raw else 'Fail'

            results[tc_id] = status

    return results


def find_next_run_col(ws, header_row: int, run_col_start: int, run_col_end: int) -> int | None:
    """Find the first Run column that is still all 'Not Run'."""
    for col in range(run_col_start, run_col_end + 1):
        all_not_run = True
        for row in range(header_row + 2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val not in (None, 'Not Run', ''):
                all_not_run = False
                break
        if all_not_run:
            return col
    return None  # All run columns are filled


def apply_result_style(cell, result: str):
    """Apply colour coding to a result cell."""
    if result == 'Pass':
        cell.fill = GREEN
        cell.font = GREEN_FONT
    elif result == 'Fail':
        cell.fill = RED
        cell.font = RED_FONT
    elif result == 'Skipped':
        cell.fill = ORANGE
        cell.font = ORANGE_FONT
    else:
        cell.fill = GREY


def update_run_tracker(wb, results: dict, run_date: str):
    """Update the Run Tracker sheet with today's results."""
    ws = wb['Run Tracker']

    # Find the next available run column
    next_col = find_next_run_col(ws, HEADER_ROW, RUN_COL_START, RUN_COL_END)
    if next_col is None:
        print('[update_tracker] WARNING: All 6 run columns are filled. No update made to Run Tracker.')
        print('                 Manually clear an old run column to record new results.')
        return

    run_label = ws.cell(row=HEADER_ROW, column=next_col).value or f'Run {next_col - RUN_COL_START + 1}'
    print(f'[update_tracker] Writing results to: {run_label} (column {next_col})')

    updated = 0
    not_found = []

    for row in range(HEADER_ROW + 2, ws.max_row + 1):
        tc_id = str(ws.cell(row=row, column=1).value or '').strip()
        if not tc_id or tc_id.startswith('▌'):
            continue

        result = results.get(tc_id)
        cell = ws.cell(row=row, column=next_col)

        if result:
            cell.value = result
            apply_result_style(cell, result)
            updated += 1
        else:
            # Test case exists in tracker but was not in this run
            cell.value = 'Not Run'
            apply_result_style(cell, 'Not Run')
            not_found.append(tc_id)

    print(f'[update_tracker] Updated {updated} test cases in Run Tracker.')
    if not_found:
        print(f'[update_tracker] {len(not_found)} TCs not found in Playwright results (marked Not Run): {not_found[:5]}{"..." if len(not_found) > 5 else ""}')


def update_test_case_register(wb, results: dict, run_date: str):
    """Update Status, Executed By, and Execution Date in the Test Case Register."""
    ws = wb['Test Case Register']

    # Column positions in Register (1-indexed):
    # A=TestID, F=Status, H=ExecutedBy, I=ExecutionDate, J=ActualResult
    STATUS_COL    = 7   # G — Status
    EXECUTED_COL  = 8   # H — Executed By
    DATE_COL      = 9   # I — Execution Date
    RESULT_COL    = 10  # J — Actual Result

    updated = 0

    for row in range(HEADER_ROW + 2, ws.max_row + 1):
        tc_id = str(ws.cell(row=row, column=1).value or '').strip()
        if not tc_id or tc_id.startswith('▌'):
            continue

        result = results.get(tc_id)
        if not result:
            continue

        # Map Playwright result to register status
        status_map = {'Pass': 'Pass', 'Fail': 'Fail', 'Skipped': 'Blocked'}
        status = status_map.get(result, 'Not Run')

        status_cell = ws.cell(row=row, column=STATUS_COL)
        status_cell.value = status
        apply_result_style(status_cell, result)

        ws.cell(row=row, column=EXECUTED_COL).value = 'Playwright (Auto)'
        ws.cell(row=row, column=DATE_COL).value      = run_date

        updated += 1

    print(f'[update_tracker] Updated {updated} rows in Test Case Register.')


def print_summary(results: dict):
    """Print a quick summary to console / CI logs."""
    passed  = sum(1 for v in results.values() if v == 'Pass')
    failed  = sum(1 for v in results.values() if v == 'Fail')
    skipped = sum(1 for v in results.values() if v == 'Skipped')
    total   = len(results)
    rate    = round((passed / total * 100), 1) if total else 0

    print('\n' + '='*50)
    print('  QA TRACKER UPDATE SUMMARY')
    print('='*50)
    print(f'  Total results received : {total}')
    print(f'  ✅ Pass                : {passed}')
    print(f'  ❌ Fail                : {failed}')
    print(f'  ⏭  Skipped             : {skipped}')
    print(f'  📊 Pass Rate           : {rate}%')
    print('='*50 + '\n')

    if failed > 0:
        failed_ids = [k for k, v in results.items() if v == 'Fail']
        print(f'  Failed TCs: {", ".join(failed_ids)}')
        print()


def main():
    parser = argparse.ArgumentParser(description='Update AMC Master Tracker from Playwright results')
    parser.add_argument('--results', required=True, help='Path to Playwright results.json')
    parser.add_argument('--excel',   required=True, help='Path to AMC_Master_Tracker.xlsx')
    args = parser.parse_args()

    run_date = date.today().strftime('%d-%m-%Y')

    print(f'[update_tracker] Reading results from: {args.results}')
    results = parse_playwright_results(args.results)
    print(f'[update_tracker] Found {len(results)} test case results.')

    print(f'[update_tracker] Loading tracker: {args.excel}')
    wb = load_workbook(args.excel)

    update_run_tracker(wb, results, run_date)
    update_test_case_register(wb, results, run_date)

    wb.save(args.excel)
    print(f'[update_tracker] Saved: {args.excel}')

    print_summary(results)


if __name__ == '__main__':
    main()